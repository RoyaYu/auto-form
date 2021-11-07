import openpyxl
from pathlib import Path

title = ["cn"]
mem_dic = {} ##“key: cn  value: list[goods:price]”

path = Path('test_file', 'test_real.xlsx')
workbook= openpyxl.load_workbook(path,data_only=True) #排表
workbook.active

# reading part
def addSheet(worksheet, index):
    """add information to mem_dic by worksheet
    sheet are sepreated by kind of goods
    goods chara name and price need to be added to mem_dic
    """
    goods_name = worksheet["b1"].value
    title.append(goods_name)
    title.append("肾"+str(index)) ## the price col title
    avg_price = worksheet["c1"].value ##均价
    #loop the data row by row
    for i in range(1, worksheet.max_row):
        diff = worksheet["a"+str(i+1)].value ##调价
        chara = worksheet["c"+str(i+1)].value ##角色
        print(chara)
        for col in worksheet.iter_cols(4, worksheet.max_column):
            cn = col[i].value
            if cn not in mem_dic:
                mem_dic[cn] = {goods_name:[[],"",0]} ##add cn to the member
            elif goods_name not in mem_dic[cn]:
                mem_dic[cn][goods_name] = [[],"",0] ##add goods to the cn
            mem_dic[cn][goods_name][2] += avg_price + diff
            if (len(mem_dic[cn][goods_name][0]) == 0):
                mem_dic[cn][goods_name][0] = [chara, 1] #create goods list
            else:
                mem_dic[cn][goods_name][0][1] += 1 ## add chara to goods list
        clean_up(mem_dic, goods_name)



def clean_up(mem_dic, goods_name):
    """ remove counting trash from mem_dic
        the first counter in goods list need to be removed
        and the None entry caused by space need to be removed
    """
    for i in enumerate(mem_dic):
        if goods_name in mem_dic[i[1]]:
            if len(mem_dic[i[1]][goods_name][0]) != 0:
                one_goods = str(mem_dic[i[1]][goods_name][0][0]) \
                 + str(mem_dic[i[1]][goods_name][0][1])
                mem_dic[i[1]][goods_name][1] += one_goods #update goods list
                mem_dic[i[1]][goods_name][0] = [] # empty the counter
    mem_dic.pop(None, None) #remove the none entry


# writing part
def write_sheet():
    """
    """
    price_sheet = workbook.create_sheet("肾表")
    for til in enumerate(title):
        price_sheet.cell(column = til[0] + 1, row = 1, value = til[1])# title
    for count in enumerate(mem_dic):
        row = count[0] + 2
        cn = count[1]
        price_sheet.cell(column = 1, row = row, value = cn) ## write cn
        personal_goods = mem_dic[cn]
        for goods_name,goods_price in personal_goods.items():
            col = title.index(goods_name) + 1
            price_sheet.cell(column = col, row = row, value = goods_price[0])
            price_sheet.cell(column = col+1, row = row, value = goods_price[1])
            ## write goods list and price
    workbook.save("output.xlsx") # save file to local


count = 0
for sheet in workbook.worksheets:
    count += 1
    addSheet(sheet,count
# remove the counter as it is no longer inneed
for i in enumerate(mem_dic):
    for j in enumerate(mem_dic[i[1]]):
        mem_dic[i[1]][j[1]].pop(0)

print(mem_dic)
write_sheet();
